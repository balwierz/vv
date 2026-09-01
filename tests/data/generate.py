#!/usr/bin/env python3
"""
Generate small fixture files for the smoke test suite.
Run from the repo root: `python3 tests/data/generate.py`.
"""
from __future__ import annotations
import gzip
import shutil
import subprocess
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    import pyarrow.ipc as ipc
except ImportError:
    sys.exit("pyarrow is required: pip install pyarrow")

# ── Parquet ──────────────────────────────────────────────────────────────────
schema = pa.schema([
    ("Chr",   pa.string()),
    ("Start", pa.int64()),
    ("End",   pa.int64()),
    ("Score", pa.float32()),
    ("Tags",  pa.list_(pa.string())),
])
def gen_rows():
    rows = []
    # 12 sorted chr1 rows then 8 sorted chr2 rows (tabix requires sorted input).
    for i in range(12):
        rows.append(("chr1", 100 + i * 1000, 200 + i * 1000))
    for i in range(8):
        rows.append(("chr2", 100 + i * 1000, 200 + i * 1000))
    return rows

_rows = gen_rows()
_tag_cycle = [["promoter"], ["enhancer", "open"], [], ["TF"], ["promoter", "TF"]]
table = pa.table({
    "Chr":   [r[0] for r in _rows],
    "Start": [r[1] for r in _rows],
    "End":   [r[2] for r in _rows],
    "Score": [round(0.05 * i, 4) for i in range(20)],
    "Tags":  [_tag_cycle[i % len(_tag_cycle)] for i in range(20)],
}, schema=schema)
# Split into 4 row groups so generic-Parquet region-query tests actually
# exercise row-group pruning (otherwise everything fits in row group 0).
pq.write_table(table, HERE / "tiny.parquet", compression="snappy", row_group_size=5)

# tiny.uint32.parquet: Start/End stored as UInt32 (a common compact encoding
# for genomic positions). The region predicate used to read only Int32/Int64
# and returned 0 for any other width, silently emptying the result.
pq.write_table(pa.table({
    "Chr":   pa.array(["chr1", "chr1", "chr1", "chr2"], pa.string()),
    "Start": pa.array([100, 1100, 2100, 500], pa.uint32()),
    "End":   pa.array([200, 1200, 2200, 600], pa.uint32()),
    "Score": pa.array([0.0, 0.1, 0.2, 0.3], pa.float64()),
}), HERE / "tiny.uint32.parquet")

# tiny.temporal.parquet: date / timestamp / decimal columns. These are
# is_numeric_type() but the numeric value extractor used to skip them, so
# --heatmap rendered them blank and --describe showed no min/max. The decimal
# min/max must honour the column scale (0.01 .. 99.99).
import datetime as _dt
import decimal as _dec
pq.write_table(pa.table({
    "n":   pa.array([1, 2, 3, 4], pa.int32()),
    "d":   pa.array([_dt.date(2024, 1, 1), _dt.date(2024, 6, 1),
                     _dt.date(2024, 12, 31), _dt.date(2025, 3, 15)], pa.date32()),
    "ts":  pa.array([_dt.datetime(2024, 1, 1, 12), _dt.datetime(2024, 6, 1, 8),
                     _dt.datetime(2024, 12, 31, 23), _dt.datetime(2025, 3, 15, 6)],
                    pa.timestamp("us")),
    "dec": pa.array([_dec.Decimal("1.50"), _dec.Decimal("2.25"),
                     _dec.Decimal("99.99"), _dec.Decimal("0.01")],
                    pa.decimal128(10, 2)),
}), HERE / "tiny.temporal.parquet")

# tiny.nested.parquet: a 2-leaf struct column sits *before* Start/End, so the
# Parquet leaf-column index of Start/End is shifted past their Arrow field
# index. Region pruning reads column statistics by leaf index; using the field
# index would read the struct's stats as "Start". meta.b is set huge (100000)
# so that wrong read makes Start.min look >= the query window end and the
# (only, matching) row group gets pruned — a region query would return 0 rows.
# With the correct leaf mapping, chr1:150-160 overlaps the row (Start 100, End
# 200) and returns 1.
_struct_t = pa.struct([("a", pa.int32()), ("b", pa.int32())])
pq.write_table(pa.table({
    "Chr":   pa.array(["chr1"], pa.string()),
    "meta":  pa.array([{"a": 0, "b": 100000}], _struct_t),
    "Start": pa.array([100], pa.int64()),
    "End":   pa.array([200], pa.int64()),
}), HERE / "tiny.nested.parquet")

# ── Arrow IPC ────────────────────────────────────────────────────────────────
with pa.OSFile(str(HERE / "tiny.arrow"), "wb") as f:
    with ipc.new_file(f, schema) as w:
        # Two batches so we exercise lazy loading.
        w.write_batch(table.slice(0, 10).to_batches()[0])
        w.write_batch(table.slice(10, 10).to_batches()[0])

# tiny.dict.arrow: an Arrow IPC file that carries dictionary (categorical)
# encoding through to the reader. The filter / stats accessors handled only
# plain String / Int / Double arrays, so a predicate on a dictionary column
# matched nothing (== / in) or every row (!= / not in), silently. The accessors
# now decode the dictionary. cat: b a b c a ; dnum (dict-int): 100 200 100 300 200
_dict_tbl = pa.table({
    "cat":  pa.array(["b", "a", "b", "c", "a"], pa.string()).dictionary_encode(),
    "dnum": pa.array([100, 200, 100, 300, 200], pa.int64()).dictionary_encode(),
    "val":  pa.array([10, 20, 30, 40, 50], pa.int64()),
})
with pa.OSFile(str(HERE / "tiny.dict.arrow"), "wb") as f:
    with ipc.new_file(f, _dict_tbl.schema) as w:
        w.write_table(_dict_tbl)

# tiny.corrupt.arrow: an Arrow IPC whose footer declares 20 record batches but
# whose batch 1 is corrupt. The file opens (schema + batch 0 are intact), so
# num_chunks() reports 20, but batches after the failed one never load. Paging
# to the last chunk read past batches_ / batch_first_row_ (a wild-pointer
# dereference — SIGSEGV in a release build, heap-buffer-overflow under ASan).
# The over-read is only reachable through the TUI / GUI (chunk_meta), so the
# regression is driven by tests/tui_corrupt_chunk_check.py. 20 batches make the
# out-of-bounds index far enough to fault reliably; the values stay small so no
# 0xFFFFFFFF byte run collides with the IPC message framing scanned below.
_cs = pa.schema([("chr", pa.string()), ("x", pa.int64())])
_tmp = HERE / "tiny.corrupt.arrow"
with pa.OSFile(str(_tmp), "wb") as f:
    with ipc.new_file(f, _cs) as w:
        for _k in range(20):
            w.write_batch(pa.record_batch({"chr": pa.array([f"chr{_k}"] * 2),
                                           "x":   pa.array([_k * 2, _k * 2 + 1],
                                                           pa.int64())}))
_data = bytearray(_tmp.read_bytes())
# Encapsulated IPC messages are each framed by a 0xFFFFFFFF continuation marker:
# [schema][batch 0][batch 1]...[EOS]. The 3rd marker starts batch 1; scramble
# 16 bytes of its flatbuffer so ReadRecordBatch(1) fails while the footer and
# batch 0 stay valid.
_marker = b"\xff\xff\xff\xff"
_pos, _i = [], 0
while True:
    _j = _data.find(_marker, _i)
    if _j < 0:
        break
    _pos.append(_j)
    _i = _j + 1
if len(_pos) >= 3:
    _b1 = _pos[2]
    for _k in range(_b1 + 8, _b1 + 8 + 16):
        _data[_k] ^= 0xFF
    _tmp.write_bytes(bytes(_data))
    # Confirm the intended shape: footer=20, batch 0 decodes, batch 1 does not.
    _r = ipc.open_file(str(_tmp))
    assert _r.num_record_batches == 20
    _r.get_batch(0)
    try:
        _r.get_batch(1)
        print("warn: tiny.corrupt.arrow batch 1 still decodes; fixture ineffective",
              file=sys.stderr)
    except Exception:
        pass
else:
    print("warn: could not locate batch 1 framing; skipping tiny.corrupt.arrow",
          file=sys.stderr)

# tiny.empty.arrow: a valid Arrow IPC with a schema but zero record batches.
# The reader seeds a zero-row batch so the schema renders, but num_chunks()
# used to report 0 (num_record_batches_) and the table view drew nothing.
with pa.OSFile(str(HERE / "tiny.empty.arrow"), "wb") as f:
    ipc.new_file(f, schema).close()

# ── BED ──────────────────────────────────────────────────────────────────────
# Use round() to avoid float32 noise leaking into the text fixture.
bed = "\n".join(
    f"{chrom}\t{start}\t{end}\tpeak_{i}\t{round(score, 4)}"
    for i, (chrom, start, end, score) in enumerate(
        zip(table["Chr"].to_pylist(),
            table["Start"].to_pylist(),
            table["End"].to_pylist(),
            table["Score"].to_pylist()))
) + "\n"
(HERE / "tiny.bed").write_text(bed)

# ── BED.gz + tabix ───────────────────────────────────────────────────────────
def have(cmd):
    return shutil.which(cmd) is not None

if have("bgzip") and have("tabix"):
    bed_gz = HERE / "tiny.bed.gz"
    if bed_gz.exists():
        bed_gz.unlink()
    # `bgzip -k` (keep input) only exists in htslib >= 1.14. Ubuntu 22.04
    # ships htslib 1.13. Stream form (`bgzip -c < src > out`) is portable.
    with open(HERE / "tiny.bed", "rb") as fin, open(bed_gz, "wb") as fout:
        subprocess.run(["bgzip", "-c"], stdin=fin, stdout=fout, check=True)
    subprocess.run(["tabix", "-f", "-p", "bed", str(bed_gz)], check=True)

    # Ensembl-named tabix BED (contigs `1` and `MT`) to exercise UCSC<->Ensembl
    # region-chromosome aliasing: a `-r chr1` / `-r chrM` query must find `1` /
    # `MT`. Sorted; MT sorts after the numeric contig.
    ens_bed = HERE / "tiny.ens.bed"
    ens_bed.write_text("1\t100\t200\tp1\n1\t500\t800\tp2\nMT\t10\t20\tmt1\n")
    ens_gz = HERE / "tiny.ens.bed.gz"
    if ens_gz.exists():
        ens_gz.unlink()
    with open(ens_bed, "rb") as fin, open(ens_gz, "wb") as fout:
        subprocess.run(["bgzip", "-c"], stdin=fin, stdout=fout, check=True)
    subprocess.run(["tabix", "-f", "-p", "bed", str(ens_gz)], check=True)
    ens_bed.unlink()   # keep only the bgzipped + indexed form
else:
    print("warn: bgzip/tabix not found; skipping tiny.bed.gz", file=sys.stderr)

# ── VCF ──────────────────────────────────────────────────────────────────────
vcf = """##fileformat=VCFv4.2
##INFO=<ID=AF,Number=1,Type=Float,Description="Allele frequency">
#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO
chr1\t100\trs1\tA\tG\t30\tPASS\tAF=0.5
chr1\t500\t.\tC\tT\t40\tPASS\tAF=0.1
chr1\t1500\t.\tG\tA\t50\tPASS\tAF=0.3
chr2\t200\t.\tT\tC\t35\tPASS\tAF=0.2
"""
(HERE / "tiny.vcf").write_text(vcf)
if have("bgzip") and have("tabix"):
    vcf_gz = HERE / "tiny.vcf.gz"
    if vcf_gz.exists():
        vcf_gz.unlink()
    with open(HERE / "tiny.vcf", "rb") as fin, open(vcf_gz, "wb") as fout:
        subprocess.run(["bgzip", "-c"], stdin=fin, stdout=fout, check=True)
    subprocess.run(["tabix", "-f", "-p", "vcf", str(vcf_gz)], check=True)

# ── FASTA ────────────────────────────────────────────────────────────────────
fa = """>seq1 chromosome 1
ACGTACGTACGTACGT
ACGTACGTACGTACGT
>seq2
GGGCCCAAATTT
>seq3 mitochondrial
TTTAAAACCCCGGGGAAAA
"""
(HERE / "tiny.fa").write_text(fa)
with open(HERE / "tiny.fa", "rb") as src, gzip.open(HERE / "tiny.fa.gz", "wb") as dst:
    shutil.copyfileobj(src, dst)

# ── 2bit (UCSC sequence container, hand-rolled) ─────────────────────────────
# Minimal 32-bit little-endian 2bit file with three short sequences.
# Format reference: https://genome.ucsc.edu/FAQ/FAQformat.html#format7
#
# Bases pack into 2 bits: T=0, C=1, A=2, G=3 (MSB-first within each byte).
import struct
def _encode_2bit(seq: str) -> tuple[bytes, list[tuple[int,int]]]:
    # Returns (packed_dna, n_blocks). N-blocks are contiguous runs of Ns;
    # the packed bytes for those bases use any value (we use T=0 fill).
    code = {"T": 0, "C": 1, "A": 2, "G": 3, "t": 0, "c": 1, "a": 2, "g": 3}
    n_blocks = []
    nrun_start = None
    cleaned = []
    for i, b in enumerate(seq):
        if b in ("N", "n"):
            cleaned.append("T")
            if nrun_start is None: nrun_start = i
        else:
            cleaned.append(b)
            if nrun_start is not None:
                n_blocks.append((nrun_start, i - nrun_start))
                nrun_start = None
    if nrun_start is not None:
        n_blocks.append((nrun_start, len(seq) - nrun_start))
    cleaned_s = "".join(cleaned)
    out = bytearray()
    for i in range(0, len(cleaned_s), 4):
        v = 0
        for j in range(4):
            v <<= 2
            if i + j < len(cleaned_s):
                v |= code[cleaned_s[i + j]]
        out.append(v)
    return bytes(out), n_blocks

_seqs = [
    ("chr1", "ACGTACGTACGTACGTACGT"),                # 20 bp, no Ns
    ("chr2", "ACGTNNNACGTACGTACGTACGTACGT"),         # 27 bp, one N-run
    ("chrM", "GGGCCCAAATTT"),                        # 12 bp
]
# Compute layout: header (16) + index entries (1+name+4 each) +
# per-seq seqRecords. We need the offsets before writing — compute
# in two passes.
index_bytes = sum(1 + len(n) + 4 for n, _ in _seqs)
header_size = 16
cursor = header_size + index_bytes
seq_records = []
offsets = []
for _name, _seq in _seqs:
    offsets.append(cursor)
    dna, nb = _encode_2bit(_seq)
    rec = struct.pack("<I", len(_seq))                    # dnaSize
    rec += struct.pack("<I", len(nb))                     # nBlockCount
    rec += b"".join(struct.pack("<I", s) for s, _ in nb)  # nBlockStarts
    rec += b"".join(struct.pack("<I", l) for _, l in nb)  # nBlockSizes
    rec += struct.pack("<I", 0)                           # maskBlockCount
    rec += struct.pack("<I", 0)                           # reserved
    rec += dna
    seq_records.append(rec)
    cursor += len(rec)

out = bytearray()
out += struct.pack("<IIII", 0x1A412743, 0, len(_seqs), 0)
for (name, _), off in zip(_seqs, offsets):
    nb = name.encode()
    out.append(len(nb))
    out += nb
    out += struct.pack("<I", off)
for rec in seq_records:
    out += rec
(HERE / "tiny.2bit").write_bytes(bytes(out))

# tiny.malformed.2bit: a header declaring 0xFFFFFFFF sequences over a 16-byte
# (header-only) file. The reader used to reserve() ~170 GB for the index up
# front (abort/OOM on systems without memory overcommit); it must now reject
# the impossible count cleanly.
(HERE / "tiny.malformed.2bit").write_bytes(
    struct.pack("<IIII", 0x1A412743, 0, 0xFFFFFFFF, 0))

# ── FASTQ ────────────────────────────────────────────────────────────────────
fq = """@read1 lane=1
ACGTACGTACGTACGT
+
IIIIIIIIIIIIIIII
@read2
TTTTAAAA
+
HHHHGGGG
@read3 short
GGG
+
III
"""
(HERE / "tiny.fq").write_text(fq)
with open(HERE / "tiny.fq", "rb") as src, gzip.open(HERE / "tiny.fq.gz", "wb") as dst:
    shutil.copyfileobj(src, dst)

# ── PAF (minimap2 pairwise alignments) ──────────────────────────────────────
paf = (
    "read1\t1000\t100\t900\t+\tchr1\t250000000\t5000\t5800\t750\t800\t60\tNM:i:50\n"
    "read2\t2000\t0\t2000\t-\tchr2\t300000000\t10000\t12000\t1900\t2000\t60\n"
    "read3\t500\t0\t500\t+\tchr1\t250000000\t30000\t30500\t480\t500\t30\tNM:i:20\tms:i:480\n"
)
(HERE / "tiny.paf").write_text(paf)
with open(HERE / "tiny.paf", "rb") as src, gzip.open(HERE / "tiny.paf.gz", "wb") as dst:
    shutil.copyfileobj(src, dst)

# ── LociSSD (Parquet + lociSSD_manifest in file-level KV metadata) ──────────
# Reference: /home/piotr/Sources/LociSSD/FORMAT_SPEC.md
# Only the manifest's presence matters for vv detection; we hand-craft a
# minimal valid manifest here instead of pulling in the full lociSSD package.
import json
loc = pa.table({
    "Chromosome":  pa.array(["chr1","chr1","chr1","chr2","chr2"], pa.string()),
    "Start":       pa.array([100, 500, 1000, 200, 1500], pa.int32()),
    "End":         pa.array([200, 800, 1200, 400, 1800], pa.int32()),
    "Name":        pa.array(["peak_0","peak_1","peak_2","peak_3","peak_4"], pa.string()),
    "Score":       pa.array([0.5, 0.7, 0.2, 0.9, 0.1], pa.float64()),
    "MaxEndSoFar": pa.array([200, 800, 1200, 400, 1800], pa.int32()),
})
manifest = json.dumps({
    "format_version": 2,
    "writer_version": "vv tests/data/generate.py",
    "created_utc":    "2026-05-07T00:00:00+00:00",
    "row_count":      5,
    "assembly":       "hg38",   # species omitted -> derived (Homo sapiens) in the banner
    "chromosomes": [
        {"name": "chr1", "rows": 3, "row_offset": 0,
         "min_start": 100,  "max_start": 1000, "max_end": 1200},
        {"name": "chr2", "rows": 2, "row_offset": 3,
         "min_start": 200,  "max_start": 1500, "max_end": 1800},
    ],
    "sort_keys": ["Chromosome", "Start", "End"],
    "default_compression": ["zstd", 3],
    "coord_dtype": "int32",
})
loc = loc.replace_schema_metadata({"lociSSD_manifest": manifest})
# Force 2 row groups (rows 0..1 and 2..4) so vv's row-group statistics
# pruning is actually exercised by region-query tests.
pq.write_table(loc, HERE / "tiny.lociss", compression="zstd", row_group_size=2)

# tiny.badcoord.lociss: a LociSSD file with a NULL in the (nullable) int Start
# column. --validate's per-row scan read each coordinate via std::stoll on the
# rendered cell; a null renders as the null symbol, which stoll cannot parse, so
# it threw std::invalid_argument and aborted the process. The columns are all the
# right (integer/string) types, so every schema check passes first — the crash
# hit on exactly the malformed file --validate exists to diagnose. Must now be a
# reported failure, not an abort.
badcoord = pa.table({
    "Chromosome":  pa.array(["chr1", "chr1", "chr1"], pa.string()),
    "Start":       pa.array([100, None, 300], pa.int32()),   # null coordinate
    "End":         pa.array([200, 250, 400], pa.int32()),
    "MaxEndSoFar": pa.array([200, 250, 400], pa.int32()),
})
_bc_manifest = json.dumps({
    "format_version": 2, "row_count": 3, "assembly": "hg38",
    "chromosomes": [{"name": "chr1", "rows": 3, "row_offset": 0}],
    "sort_keys": ["Chromosome", "Start", "End"], "coord_dtype": "int32",
})
badcoord = badcoord.replace_schema_metadata({"lociSSD_manifest": _bc_manifest})
pq.write_table(badcoord, HERE / "tiny.badcoord.lociss", compression="zstd")

# ── LociSSD v4 "colblock" (tiny.v4.lociss + .idx) ───────────────────────────
# A minimal hand-rolled v4 fixture (only pyarrow for the zstd frames). 2
# chromosomes, block_rows=3 → 2 blocks (a block never spans a chromosome),
# exercising codecs DELTA(Start)/LENGTH(End)/DICT(Strand)/FRONTCODE(ID)/RAW(Count)
# /BOOL(Flag), a NULL (Count row 1) and an empty string (ID row 1, distinct from
# null). Spec: Loci1/docs/lociss_columnar_format_spec.md. Lets the v4 reader be
# tested without the private `loci` package.
def _v4_zstd(b):
    return pa.compress(memoryview(b), codec="zstd", asbytes=True)
def _v4_chunk(payload, n, valid=None):
    if valid is None:
        body = b"\x00" + payload
    else:
        bm = bytearray((n + 7) // 8)
        for i, v in enumerate(valid):
            if v: bm[i >> 3] |= 1 << (i & 7)
        body = b"\x01" + bytes(bm) + payload
    return _v4_zstd(body)
def _v4_i32(vals):  # RAW int32 (placeholder 0 for null positions)
    return b"".join(struct.pack("<i", v if v is not None else 0) for v in vals)
def _v4_delta(vals):
    out = bytearray(); prev = 0
    for i, v in enumerate(vals):
        out += struct.pack("<i", v - (prev if i else 0)); prev = v
    return bytes(out)
def _v4_length(starts, ends):
    return b"".join(struct.pack("<i", e - s) for s, e in zip(starts, ends))
def _v4_dict(vals):
    uniq, idx = [], {}
    for v in vals:
        if v not in idx: idx[v] = len(uniq); uniq.append(v)
    blob = b"".join(u.encode() for u in uniq)
    offs = [0]
    for u in uniq: offs.append(offs[-1] + len(u.encode()))
    out = struct.pack("<I", len(uniq)) + b"".join(struct.pack("<I", o) for o in offs) + blob
    out += b"".join(struct.pack("<B", idx[v]) for v in vals)  # n_dict<=256 -> u8 codes
    return out
def _v4_frontcode(vals):
    lcp, slen, suf, prev = [], [], bytearray(), b""
    for v in vals:
        vb = v.encode(); l = 0
        while l < min(len(prev), len(vb)) and prev[l] == vb[l]: l += 1
        lcp.append(l); slen.append(len(vb) - l); suf += vb[l:]; prev = vb
    return (b"".join(struct.pack("<I", x) for x in lcp) +
            b"".join(struct.pack("<I", x) for x in slen) + bytes(suf))
def _v4_bool(vals):
    bm = bytearray((len(vals) + 7) // 8)
    for i, v in enumerate(vals):
        if v: bm[i >> 3] |= 1 << (i & 7)
    return bytes(bm)

_v4_blocks = [
    (0, {"Start": [100, 150, 200], "End": [101, 152, 201],
         "Strand": ["+", "-", "+"], "ID": ["rs1", "", "rs10"],
         "Count": [10, None, 30], "Flag": [True, False, True]}),
    (1, {"Start": [300, 350], "End": [301, 360],
         "Strand": ["-", "+"], "ID": ["rs100", "rs101"],
         "Count": [40, 50], "Flag": [False, True]}),
]
_v4_stored = ["Start", "End", "Strand", "ID", "Count", "Flag"]
_v4_schema = {"Start": "int32", "End": "int32", "Strand": "utf8",
              "ID": "utf8", "Count": "int32", "Flag": "bool"}
_v4_codecs = {"Start": 1, "End": 2, "Strand": 3, "ID": 4, "Count": 0, "Flag": 6}

_data = bytearray(b"LSB1" + bytes([4]) + b"\x00\x00\x00")
_off, _clen, _cids, _mins, _maxs, _nr, _pmax = [], [], [], [], [], [], []
_run = {}
for _cid, _rows in _v4_blocks:
    _n = len(_rows["Start"])
    _cids.append(_cid); _nr.append(_n)
    _mins.append(min(_rows["Start"])); _me = max(_rows["End"]); _maxs.append(_me)
    _run[_cid] = max(_run.get(_cid, -(1 << 62)), _me); _pmax.append(_run[_cid])
    for _col in _v4_stored:
        _vals = _rows[_col]
        _valid = [v is not None for v in _vals] if any(v is None for v in _vals) else None
        _cc = _v4_codecs[_col]
        if   _cc == 0: _pay = _v4_i32(_vals)
        elif _cc == 1: _pay = _v4_delta(_vals)
        elif _cc == 2: _pay = _v4_length(_rows["Start"], _vals)
        elif _cc == 3: _pay = _v4_dict([v if v is not None else "" for v in _vals])
        elif _cc == 4: _pay = _v4_frontcode([v if v is not None else "" for v in _vals])
        elif _cc == 6: _pay = _v4_bool([bool(v) for v in _vals])
        _ck = _v4_chunk(_pay, _n, _valid)
        _off.append(len(_data)); _clen.append(len(_ck)); _data += _ck

_meta = {"format_version": 4, "writer_version": "vv tests/data/generate.py",
         "stored": _v4_stored, "schema": _v4_schema, "codecs": _v4_codecs,
         "coord_dtype": "int32", "block_rows": 3, "row_count": sum(_nr),
         "rank_to_name": {"0": "chr1", "1": "chr2"},
         "assembly": "hg38", "species": None}
_mb = json.dumps(_meta).encode()
_idx = bytearray(b"LSI1" + bytes([1]) + b"\x00\x00\x00")
_idx += struct.pack("<IIII", len(_v4_blocks), len(_v4_stored), 0, len(_mb)) + _mb
_idx += b"".join(struct.pack("<i", x) for x in _cids)
_idx += b"".join(struct.pack("<i", x) for x in _mins)
_idx += b"".join(struct.pack("<i", x) for x in _maxs)
_idx += b"".join(struct.pack("<I", x) for x in _nr)
_idx += b"".join(struct.pack("<i", x) for x in _pmax)
_idx += b"".join(struct.pack("<Q", x) for x in _off)
_idx += b"".join(struct.pack("<I", x) for x in _clen)
(HERE / "tiny.v4.lociss").write_bytes(bytes(_data))
(HERE / "tiny.v4.lociss.idx").write_bytes(bytes(_idx))

# V4.1 single-file variant (tiny.v41.lociss, no sidecar): same chunks + the LSI1
# index stored INLINE, located by a 24-byte LSIX trailer at EOF. Header offset-5
# flags byte has bit0 (INLINE_INDEX) set; the chunks keep their offsets (8+), so
# the index's col_offset pointers stay valid. (Spec §8a.)
_v41 = bytearray(b"LSB1" + bytes([4, 1, 0, 0])) + bytes(_data[8:])  # version 4, flags bit0=1
_v41_off = len(_v41)                                                # == len(_data)
_v41 += _idx
_v41 += struct.pack("<QQ", _v41_off, len(_idx)) + b"LSIX" + bytes([1, 0, 0, 0])
(HERE / "tiny.v41.lociss").write_bytes(bytes(_v41))

# ── Hostile / malformed v4 fixtures (memory-safety hardening) ────────────────
# A single-block v4 sidecar file whose one column carries a hand-crafted chunk
# body (pre-zstd: has_nulls byte + codec payload). Used to prove the reader
# rejects corrupt DICT/ARENA offsets and decompression bombs instead of reading
# out of bounds. `vv` must exit non-zero cleanly (never crash) on these.
def _v4_hostile(name, colname, type_str, codec_id, n, body):
    data = bytearray(b"LSB1" + bytes([4, 0, 0, 0]))   # v4 sidecar header, flags=0
    ck = _v4_zstd(body)
    off, clen = len(data), len(ck)
    data += ck
    meta = {"format_version": 4, "writer_version": "vv test-hostile",
            "stored": [colname], "schema": {colname: type_str},
            "codecs": {colname: codec_id}, "coord_dtype": "int32",
            "block_rows": n, "row_count": n, "rank_to_name": {"0": "chr1"},
            "assembly": "hg38", "species": None}
    mb = json.dumps(meta).encode()
    idx = bytearray(b"LSI1" + bytes([1, 0, 0, 0]))
    idx += struct.pack("<IIII", 1, 1, 0, len(mb)) + mb
    idx += struct.pack("<iiiIi", 0, 0, 1, n, 1)       # cid,min_start,max_end,n_rows,pmax
    idx += struct.pack("<Q", off) + struct.pack("<I", clen)
    (HERE / name).write_bytes(bytes(data))
    (HERE / (name + ".idx")).write_bytes(bytes(idx))

# 1) DICT with a non-monotone offset (off[1] < off[0]) → length would underflow.
_baddict = struct.pack("<I", 1) + struct.pack("<II", 5, 0) + b"\x00\x00"  # n_dict=1, off=[5,0], codes
_v4_hostile("tiny.v4.baddict.lociss", "Name", "utf8", 3, 2, b"\x00" + _baddict)
# 2) ARENA with an offset past the blob (o1=100 > 3-byte blob).
_badarena = struct.pack("<III", 0, 100, 100) + b"abc"                     # off[n+1]=off[3]
_v4_hostile("tiny.v4.badarena.lociss", "Name", "utf8", 5, 2, b"\x00" + _badarena)
# 3) Valid single-string column with NO Start/End — a -r query must error, not OOB.
_okdict = struct.pack("<I", 1) + struct.pack("<II", 0, 3) + b"abc" + b"\x00\x00"
_v4_hostile("tiny.v4.nocoord.lociss", "Name", "utf8", 3, 2, b"\x00" + _okdict)
# 5) Start / End declared as a non-int type (int8). decode_colblock honours the
#    declared schema type, but the region-overlap path cast the decoded array to
#    Int32Array unconditionally, reading 4 bytes per 1-byte element — a heap
#    over-read past the padded Arrow buffer once the column is wide enough (n=64
#    here). Must be rejected at open. Two columns, so it can't use _v4_hostile.
def _v4_badcoordtype():
    def _raw(payload):
        return _v4_zstd(b"\x00" + payload)   # has_nulls=0, then raw int8 payload
    _n = 64
    _sc = _raw(bytes([(10 + i) & 0x7f for i in range(_n)]))
    _ec = _raw(bytes([(50 + i) & 0x7f for i in range(_n)]))
    _data = bytearray(b"LSB1" + bytes([4, 0, 0, 0]))
    _os = len(_data); _data += _sc
    _oe = len(_data); _data += _ec
    _meta = {"format_version": 4, "writer_version": "vv test-hostile",
             "stored": ["Start", "End"], "schema": {"Start": "int8", "End": "int8"},
             "codecs": {"Start": 0, "End": 0}, "coord_dtype": "int32",
             "block_rows": _n, "row_count": _n, "rank_to_name": {"0": "chr1"},
             "assembly": "hg38", "species": None}
    _mb = json.dumps(_meta).encode()
    _idx = bytearray(b"LSI1" + bytes([1, 0, 0, 0]))
    _idx += struct.pack("<IIII", 1, 2, 0, len(_mb)) + _mb          # n_blocks=1, n_cols=2
    _idx += struct.pack("<i", 0) + struct.pack("<i", 0) + struct.pack("<i", 200)
    _idx += struct.pack("<I", _n) + struct.pack("<i", 200)         # n_rows, prefix_max_end
    _idx += struct.pack("<QQ", _os, _oe)                          # col_offset (SoA)
    _idx += struct.pack("<II", len(_sc), len(_ec))                # col_clen  (SoA)
    (HERE / "tiny.v4.badcoordtype.lociss").write_bytes(bytes(_data))
    (HERE / "tiny.v4.badcoordtype.lociss.idx").write_bytes(bytes(_idx))
_v4_badcoordtype()

# 4) zstd decompression bomb: a tiny compressed chunk inflating past the ~64 MiB
#    cap (n=1 → max_out ≈ 64 MiB). Body is 65 MiB of zeros → a few bytes zstd'd.
_v4_hostile("tiny.v4.zbomb.lociss", "V", "int8", 0, 1, bytes(65 << 20))

# ── SQLite (two tables: peaks + samples) ────────────────────────────────────
import sqlite3
sqlite_path = HERE / "tiny.sqlite"
if sqlite_path.exists():
    sqlite_path.unlink()
con = sqlite3.connect(sqlite_path)
cur = con.cursor()
cur.execute("""CREATE TABLE peaks(
    chrom TEXT NOT NULL, start INTEGER, end INTEGER,
    score REAL, name TEXT)""")
cur.executemany("INSERT INTO peaks VALUES(?,?,?,?,?)", [
    ("chr1", 100, 200, 5.2, "p1"),
    ("chr1", 500, 800, 8.1, "p2"),
    ("chr2", 1000, 1300, 3.4, "p3"),
])
cur.execute("""CREATE TABLE samples(
    sample_id INTEGER PRIMARY KEY, name TEXT, depth REAL)""")
cur.executemany("INSERT INTO samples VALUES(?,?,?)", [
    (1, "sampleA", 12.5),
    (2, "sampleB", 8.7),
    (3, "sampleC", 15.1),
])
con.commit()
con.close()

# tiny.types.sqlite: exercises NUMERIC-affinity columns (DATE / NUMERIC /
# DATETIME / BOOLEAN). These used to be mapped to Arrow double and read via
# sqlite3_column_double, which corrupted text dates ('2026-06-10' → 2026.0)
# and rounded 64-bit integers beyond a double's 2^53 exact range.
types_path = HERE / "tiny.types.sqlite"
if types_path.exists():
    types_path.unlink()
con = sqlite3.connect(types_path)
cur = con.cursor()
cur.execute("""CREATE TABLE events(
    d DATE, ts DATETIME, big NUMERIC, flag BOOLEAN, label TEXT)""")
cur.executemany("INSERT INTO events VALUES(?,?,?,?,?)", [
    # 9007199254740993 == 2**53 + 1, NOT exactly representable as a double.
    ("2026-06-10", "2026-06-10 12:34:56", 9007199254740993, 1, "alpha"),
    ("2025-01-02", "2025-01-02 00:00:00", 9007199254740995, 0, "beta"),
])
con.commit()
con.close()

# tiny.quoteid.sqlite: a table whose name contains a double quote (a"b). vv
# quotes the table name when building its PRAGMA / SELECT, so without escaping
# the embedded quote it produced the malformed (and, for an untrusted DB,
# injectable) SQL `"a"b"` and failed to open the table.
quote_path = HERE / "tiny.quoteid.sqlite"
if quote_path.exists():
    quote_path.unlink()
con = sqlite3.connect(quote_path)
cur = con.cursor()
cur.execute('CREATE TABLE "a""b" (id INTEGER, label TEXT)')
cur.executemany('INSERT INTO "a""b" VALUES(?,?)', [(1, "hello"), (2, "world")])
con.commit()
con.close()

# ── GTF (gencode-shaped, for --expand attributes) ────────────────────────────
# No .gff/.gtf fixture existed. Shaped like real gencode: GTF-style
# `key "value";` attributes, a repeated `tag` key (gencode repeats it), and a
# key that appears only on LATER rows (`exon_number`) so the documented
# "-n preview and a full scan can disagree on the schema" case is testable.
with open(HERE / "tiny.gtf", "w") as f:
    f.write("##description: evidence-based test fixture\n")
    f.write("##provider: GENCODE\n")
    rows = [
        ("gene",       11869, 14409,
         'gene_id "ENSG00000223972.5"; gene_type "pseudogene"; '
         'gene_name "DDX11L1"; level 2; tag "basic"; tag "Ensembl_canonical";'),
        ("transcript", 11869, 14409,
         'gene_id "ENSG00000223972.5"; transcript_id "ENST00000456328.2"; '
         'gene_name "DDX11L1"; transcript_support_level "1"; tag "basic";'),
        ("exon",       11869, 12227,
         'gene_id "ENSG00000223972.5"; transcript_id "ENST00000456328.2"; '
         'exon_number 1; gene_name "DDX11L1";'),
        ("exon",       12613, 12721,
         'gene_id "ENSG00000223972.5"; transcript_id "ENST00000456328.2"; '
         'exon_number 2; gene_name "DDX11L1";'),
    ]
    for feat, beg, end, attrs in rows:
        f.write("chr1\tHAVANA\t%s\t%d\t%d\t.\t+\t.\t%s\n" % (feat, beg, end, attrs))

# ── BAM (for --pileup tests; needs pysam) ────────────────────────────────────
# Generate a sorted + indexed BAM with three reads on chr1:100 covering
# positions 100–119, so the pileup engine emits 20 rows when invoked.
try:
    import pysam                                              # type: ignore
except ImportError:
    print("warn: pysam not found; skipping tiny.bam", file=sys.stderr)
else:
    bam_path = HERE / "tiny.bam"
    if bam_path.exists():
        bam_path.unlink()
    bai_path = HERE / "tiny.bam.bai"
    if bai_path.exists():
        bai_path.unlink()
    bam_header = {
        "HD": {"VN": "1.6", "SO": "coordinate"},
        "SQ": [
            {"SN": "chr1", "LN": 1000},
            {"SN": "chr2", "LN": 1000},
        ],
    }
    # 3 reads spanning chr1:100-119 with deliberate mismatches at pos 105
    # so the decoded pileup shows non-zero off-ref counts.
    reads = [
        # forward strand, 20bp, all match-T except pos 105 (offset 5) = G
        ("r1", 99,  "TTTTTGTTTTTTTTTTTTTT", 0,    True),
        # reverse strand, 20bp, all match-T except pos 105 = G
        ("r2", 99,  "TTTTTGTTTTTTTTTTTTTT", 16,   True),
        # forward strand starting at pos 102, only 17bp long (covers 102-118)
        ("r3", 101, "TTTGTTTTTTTTTTTTT",     0,    True),
    ]
    with pysam.AlignmentFile(str(bam_path), "wb", header=bam_header) as bf:
        for (qn, pos0, seq, flag, _) in reads:
            r = pysam.AlignedSegment(header=bf.header)
            r.query_name = qn
            r.flag = flag
            r.reference_id = 0
            r.reference_start = pos0
            r.mapping_quality = 60
            r.cigarstring = f"{len(seq)}M"
            r.query_sequence = seq
            r.query_qualities = pysam.qualitystring_to_array("I" * len(seq))
            r.next_reference_id = -1
            r.next_reference_start = -1
            r.template_length = 0
            bf.write(r)
    pysam.index(str(bam_path))

    # Reference FASTA for reference-aware pileup (`vv --pileup -f`). chr1 is all
    # T over the covered region 100-119 EXCEPT it keeps T at pos 105 — where the
    # reads carry a G — so that column shows a mismatch (G/g) while every other
    # column matches (./,). Pos 110 is lowercase 't' to exercise ref-column
    # uppercasing + case-insensitive matching. Byte-checked vs `samtools mpileup
    # -B -f` (vv does no BAQ, hence -B).
    ref_fa = HERE / "tiny.pileup.fa"
    win = "T" * 10 + "t" + "T" * 9                 # positions 100..119 (110 = 't')
    chr1_ref = "A" * 99 + win + "A" * (1000 - 119)  # 1-based; length 1000
    assert len(chr1_ref) == 1000
    with open(ref_fa, "w") as f:
        f.write(">chr1\n" + chr1_ref + "\n")
        f.write(">chr2\n" + "A" * 1000 + "\n")
    if (HERE / "tiny.pileup.fa.fai").exists():
        (HERE / "tiny.pileup.fa.fai").unlink()
    pysam.faidx(str(ref_fa))

    # tiny.cram: the same three reads as tiny.bam, CRAM-encoded against
    # tiny.pileup.fa (which was just written above, so no network reference
    # lookup is needed). Exercises `-r` on CRAM and `-f/--fasta` as the
    # reference for CRAM decoding — CRAM stores bases as differences from a
    # reference, so without one htslib falls back to $REF_PATH / $REF_CACHE.
    cram_path = HERE / "tiny.cram"
    for stale in (cram_path, HERE / "tiny.cram.crai"):
        if stale.exists():
            stale.unlink()
    pysam.view("-C", "-T", str(ref_fa), "-o", str(cram_path), str(bam_path),
               catch_stdout=False)
    pysam.index(str(cram_path))

    # tiny.splice.bam: reads with a reference skip (CIGAR N, e.g. an RNA-seq
    # intron) and a deletion (CIGAR D), on both strands, so --pileup can be
    # checked byte-for-byte against `samtools mpileup`. Refskips must render as
    # '>' / '<' (not '*'), and del/refskip quality columns use the real base
    # quality, never '*'.
    sp_path = HERE / "tiny.splice.bam"
    for p in (sp_path, HERE / "tiny.splice.bam.bai"):
        if p.exists():
            p.unlink()
    sp_reads = [
        ("sfwd", 0,  "5M3N5M", "AAAAACCCCC", 0),    # refskip, forward → '>'
        ("srev", 0,  "5M3N5M", "GGGGGTTTTT", 16),   # refskip, reverse → '<'
        ("sdel", 20, "5M3D5M", "AAAAACCCCC", 0),    # deletion → '*'
    ]
    with pysam.AlignmentFile(str(sp_path), "wb", header=bam_header) as bf:
        for (qn, pos0, cig, seq, flag) in sp_reads:
            r = pysam.AlignedSegment(header=bf.header)
            r.query_name = qn
            r.flag = flag
            r.reference_id = 0
            r.reference_start = pos0
            r.mapping_quality = 60
            r.cigarstring = cig
            r.query_sequence = seq
            r.query_qualities = pysam.qualitystring_to_array("I" * len(seq))
            r.next_reference_id = -1
            r.next_reference_start = -1
            bf.write(r)
    pysam.index(str(sp_path))

    # tiny.noseq.bam / tiny.noseqins.bam: reads with SEQ='*' (l_qseq == 0) but a
    # full CIGAR. This is legal SAM — an aligner may store the alignment without
    # the query bases — and htslib decodes it, so it reaches the pileup. The
    # pileup base is read at a CIGAR-derived query offset; with no query bases
    # that offset ran past the packed SEQ array, an out-of-bounds read of heap
    # bytes rendered as base calls, and past the allocation it faulted. The
    # 20000M read exercises the base-column read; the 1M20000I1M read exercises
    # the insertion-expansion read. The CIGAR lengths are large enough that the
    # over-read crosses the record allocation, so the ASan/UBSan job aborts on a
    # regression; a bounded reader prints 'N' for the missing bases and exits 0.
    for _name, _cigar in (("tiny.noseq.bam", "20000M"),
                          ("tiny.noseqins.bam", "1M20000I1M")):
        _p = HERE / _name
        if _p.exists():
            _p.unlink()
        with pysam.AlignmentFile(str(_p), "wb", header=bam_header) as bf:
            r = pysam.AlignedSegment(header=bf.header)
            r.query_name = "r1"
            r.flag = 0
            r.reference_id = 0
            r.reference_start = 0
            r.mapping_quality = 60
            r.cigarstring = _cigar
            r.query_sequence = None          # SEQ = '*'  -> l_qseq = 0
            r.next_reference_id = -1
            r.next_reference_start = -1
            r.template_length = 0
            bf.write(r)
        # Whole-file --pileup does not need an index, so none is written.

# ── samtools mpileup (single-sample + two-sample fixtures) ──────────────────
# Real samtools mpileup output. Six columns for single-sample; the
# two-sample variant has 3 + 3*2 = 9 columns.
# Hand-crafted to be self-consistent: depth == count of base events in the
# bases column == length of the quals column (after the decoder consumes
# `^X` start markers, `$` end markers, and `+N<seq>` / `-N<seq>` indels).
# Row 1 carries a single mismatch (C) on each strand, plus one insertion
# and one deletion attached to forward-strand reads, so the decoded view
# has non-zero values across the allele / indel columns.
mpileup_single = (
    "chr1\t100\tA\t12\t.+2AC.,,..C,c-1G,..\tIHGFGHIGFGFG\n"
    "chr1\t101\tG\t13\t..,GG.gg..,.,\tIHHHGGGHGFFFG\n"
    "chr1\t102\tT\t8\t..,..,,.\tHHHGGGFG\n"
    "chr2\t500\tC\t20\t..,,...,..,.,..,..,.\tHGFFGGGFGGFFGGFFGGFG\n"
)
(HERE / "tiny.mpileup").write_text(mpileup_single)

mpileup_multi = (
    "chr1\t100\tA\t12\t..,,..C,c,..\tIHGFGHIGFGFG\t5\t..,.,\tHGFGF\n"
    "chr1\t101\tG\t13\t..,GG.gg..,.,\tIHHHGGGHGFFFG\t4\t...G\tHGFG\n"
    "chr2\t500\tC\t20\t..,,...,..,.,..,..,.\tHGFFGGGFGGFFGGFFGGFG\t7\t..,..,.\tGFGFGFG\n"
)
(HERE / "tiny.multi.mpileup").write_text(mpileup_multi)

# Bgzip + tabix index the single-sample file so range queries are exercised.
# Mpileup uses single-position records, so tabix gets `-s 1 -b 2 -e 2`.
if have("bgzip") and have("tabix"):
    mp_gz = HERE / "tiny.mpileup.gz"
    if mp_gz.exists():
        mp_gz.unlink()
    with open(HERE / "tiny.mpileup", "rb") as fin, open(mp_gz, "wb") as fout:
        subprocess.run(["bgzip", "-c"], stdin=fin, stdout=fout, check=True)
    subprocess.run(["tabix", "-f", "-s", "1", "-b", "2", "-e", "2", str(mp_gz)],
                   check=True)

# ── OpenDocument Spreadsheet (.ods, two sheets, requires odfpy) ──────────────
try:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
except ImportError:
    print("warn: odfpy not found; skipping tiny.ods", file=sys.stderr)
else:
    ods_path = HERE / "tiny.ods"
    if ods_path.exists():
        ods_path.unlink()
    ods = OpenDocumentSpreadsheet()

    def _add_sheet(name, rows):
        t = Table(name=name)
        for row in rows:
            tr = TableRow()
            for cell in row:
                # Distinguish strings from numerics so the typed value
                # attributes get set (lets Arrow CSV infer correctly).
                if isinstance(cell, str):
                    tc = TableCell(valuetype="string")
                else:
                    tc = TableCell(valuetype="float", value=str(cell))
                tc.addElement(P(text=str(cell)))
                tr.addElement(tc)
            t.addElement(tr)
        ods.spreadsheet.addElement(t)

    _add_sheet("peaks", [
        ["chrom", "start", "end", "score", "name"],
        ["chr1",  100,     200,   5.2,     "p1"],
        ["chr1",  500,     800,   8.1,     "p2"],
        ["chr2",  1000,    1300,  3.4,     "p3"],
    ])
    _add_sheet("samples", [
        ["sample_id", "name", "depth"],
        [1, "sampleA", 12.5],
        [2, "sampleB", 8.7],
        [3, "sampleC", 15.1],
    ])
    ods.save(str(ods_path))

# ── Ragged ODS: a row wider than the header (.ods) ───────────────────────────
# A 3-column header followed by a 4-column data row. The reader used to lock the
# width to the header and Arrow's CSV reader then rejected the whole sheet
# ("Expected 3 columns, got 4"). The fix pads every row to the widest, naming
# the header's overflow column "col4".
try:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
except ImportError:
    print("warn: odfpy not found; skipping tiny.ragged.ods", file=sys.stderr)
else:
    rg_path = HERE / "tiny.ragged.ods"
    if rg_path.exists():
        rg_path.unlink()
    doc = OpenDocumentSpreadsheet()
    t = Table(name="ragged")
    for row in [["chrom", "start", "end"],
                ["chr1", 100, 200],
                ["chr2", 300, 400, "EXTRA"],   # wider than the header
                ["chr3", 500]]:                # shorter than the header
        tr = TableRow()
        for cell in row:
            tc = (TableCell(valuetype="string") if isinstance(cell, str)
                  else TableCell(valuetype="float", value=str(cell)))
            tc.addElement(P(text=str(cell)))
            tr.addElement(tc)
        t.addElement(tr)
    doc.spreadsheet.addElement(t)
    doc.save(str(rg_path))

    # tiny.merged.ods: a horizontally-merged (column-spanned) cell. The value
    # sits in the merge's top-left cell; the spanned columns are filled by
    # table:covered-table-cell elements. The reader dropped those covered cells,
    # so every column after a merge shifted left — here "x"/"y" landed under
    # b/c instead of c/d. The merge is on the second row: [M spans a+b], then
    # x under c and y under d, so the aligned row is M, "", x, y.
    from odf.table import CoveredTableCell        # type: ignore
    mg_path = HERE / "tiny.merged.ods"
    if mg_path.exists():
        mg_path.unlink()
    mdoc = OpenDocumentSpreadsheet()
    mt = Table(name="merged")
    hr = TableRow()
    for h in ("a", "b", "c", "d"):
        hc = TableCell(valuetype="string"); hc.addElement(P(text=h))
        hr.addElement(hc)
    mt.addElement(hr)
    dr = TableRow()
    mc = TableCell(valuetype="string", numbercolumnsspanned=2,
                   numberrowsspanned=1)
    mc.addElement(P(text="M")); dr.addElement(mc)
    dr.addElement(CoveredTableCell())            # the covered grid position
    for val in ("x", "y"):
        c = TableCell(valuetype="string"); c.addElement(P(text=val))
        dr.addElement(c)
    mt.addElement(dr)
    mdoc.spreadsheet.addElement(mt)
    mdoc.save(str(mg_path))

    # tiny.rowrep.ods: a *non-empty* row carrying table:number-rows-repeated.
    # The reader used to ignore the attribute and emit the row once, silently
    # dropping the duplicates. The "dup" row (repeat=3) must yield 3 rows, so
    # the sheet has 5 data rows total (a + dup×3 + z) under a 2-column header.
    rr_path = HERE / "tiny.rowrep.ods"
    if rr_path.exists():
        rr_path.unlink()
    rdoc = OpenDocumentSpreadsheet()
    rt = Table(name="rep")

    def _rr_cell(c):
        tc = (TableCell(valuetype="string") if isinstance(c, str)
              else TableCell(valuetype="float", value=str(c)))
        tc.addElement(P(text=str(c)))
        return tc

    def _rr_row(cells, repeat=None):
        tr = TableRow(numberrowsrepeated=str(repeat)) if repeat else TableRow()
        for c in cells:
            tr.addElement(_rr_cell(c))
        rt.addElement(tr)

    _rr_row(["k", "v"])                # header
    _rr_row(["a", 1])
    _rr_row(["dup", 9], repeat=3)      # non-empty row repeated 3×
    _rr_row(["z", 2])
    rdoc.spreadsheet.addElement(rt)
    rdoc.save(str(rr_path))

# ── HDF5 / AnnData fixtures ──────────────────────────────────────────────────
# tiny.h5ad: small AnnData with a CSR-sparse X, two obs columns
# (one categorical), one var column, one obsm embedding. Exercises
# the AnnData layout parser (summary + obs + var + X-preview + obsm tabs).
# tiny.h5: plain HDF5 with nested groups + a 1-D and a 2-D dataset to
# exercise the generic hierarchy walker + dataset readers.
try:
    import h5py                                              # type: ignore
    import numpy as np                                       # type: ignore
except ImportError:
    print("warn: h5py / numpy not found; skipping tiny.h5 / tiny.h5ad",
          file=sys.stderr)
else:
    # Plain HDF5 file with a group hierarchy and two datasets.
    h5_path = HERE / "tiny.h5"
    if h5_path.exists():
        h5_path.unlink()
    with h5py.File(h5_path, "w") as f:
        grp = f.create_group("counts")
        grp.create_dataset("rows", data=np.arange(10, dtype=np.int64))
        grp.create_dataset("matrix",
                            data=np.arange(20, dtype=np.float64).reshape(10, 2))
        meta = f.create_group("meta")
        meta.attrs["created_by"] = "vv tests"
        meta.create_dataset("labels",
                              data=np.array(["a", "b", "c"], dtype="S2"))

    # tiny.big1d.h5: a generic HDF5 whose 1-D dataset (1500 elements) exceeds
    # the 1000-row preview cap, so the Dataset1D tab must render "first 1000 of
    # 1500 rows" instead of reading the whole array into RAM.
    big1d_path = HERE / "tiny.big1d.h5"
    if big1d_path.exists():
        big1d_path.unlink()
    with h5py.File(big1d_path, "w") as f:
        f.create_dataset("big", data=np.arange(1500, dtype=np.int64))

    # tiny.wideenum.h5: an enum column whose BASE integer type is 32 bytes
    # wide. H5Tget_member_value writes one value in the base type, so reading
    # a member into a bare int64 wrote 24 bytes past it. The values are 0/1/2,
    # whose low 8 little-endian bytes are correct, so the RENDERED OUTPUT IS
    # THE SAME with and without the fix — this fixture discriminates only
    # under the ASan/UBSan CI job, where the unfixed code aborts with
    # stack-buffer-overflow inside H5Tget_member_value.
    #
    # h5py cannot express this: enum_insert() takes a C long long, so it
    # cannot write a value wider than 8 bytes, and its low-level dataset
    # writer produces a type h5dump itself rejects. Drive libhdf5 through
    # ctypes instead, mirroring the C API call sequence exactly.
    wideenum_path = HERE / "tiny.wideenum.h5"
    try:
        import ctypes, ctypes.util
        _lib = ctypes.CDLL(ctypes.util.find_library("hdf5") or "libhdf5.so")
        _lib.H5open()          # populates the predefined-type globals below
        _hid = ctypes.c_long
        for _fn, _arg, _res in (
            ("H5Fcreate", [ctypes.c_char_p, ctypes.c_uint, _hid, _hid], _hid),
            ("H5Tcopy", [_hid], _hid),
            ("H5Tset_size", [_hid, ctypes.c_size_t], ctypes.c_int),
            ("H5Tenum_create", [_hid], _hid),
            ("H5Tenum_insert", [_hid, ctypes.c_char_p, ctypes.c_void_p], ctypes.c_int),
            ("H5Screate_simple",
             [ctypes.c_int, ctypes.POINTER(ctypes.c_ulonglong), ctypes.c_void_p], _hid),
            ("H5Dcreate2", [_hid, ctypes.c_char_p, _hid, _hid, _hid, _hid, _hid], _hid),
            ("H5Dwrite", [_hid, _hid, _hid, _hid, _hid, ctypes.c_void_p], ctypes.c_int),
            ("H5Dclose", [_hid], ctypes.c_int), ("H5Sclose", [_hid], ctypes.c_int),
            ("H5Tclose", [_hid], ctypes.c_int), ("H5Fclose", [_hid], ctypes.c_int),
        ):
            _f = getattr(_lib, _fn); _f.argtypes = _arg; _f.restype = _res
        if wideenum_path.exists():
            wideenum_path.unlink()
        _W = 32                                  # enum base width, in bytes
        _fid = _lib.H5Fcreate(str(wideenum_path).encode(), 0x0002, 0, 0)
        if _fid < 0:
            raise RuntimeError("H5Fcreate failed")
        _base = _lib.H5Tcopy(_hid.in_dll(_lib, "H5T_STD_I64LE_g").value)
        _lib.H5Tset_size(_base, _W)
        _et = _lib.H5Tenum_create(_base)
        for _name, _v in ((b"LOW", 0), (b"MID", 1), (b"HIGH", 2)):
            _buf = (ctypes.c_ubyte * _W)(); _buf[0] = _v      # little-endian
            _lib.H5Tenum_insert(_et, _name, ctypes.byref(_buf))
        _dims = (ctypes.c_ulonglong * 1)(3)
        _sid = _lib.H5Screate_simple(1, _dims, None)
        _did = _lib.H5Dcreate2(_fid, b"grade", _et, _sid, 0, 0, 0)
        _data = (ctypes.c_ubyte * (3 * _W))()
        for _i, _v in enumerate((0, 1, 2)):
            _data[_i * _W] = _v
        _lib.H5Dwrite(_did, _et, 0, 0, 0, ctypes.byref(_data))
        for _close, _obj in ((_lib.H5Dclose, _did), (_lib.H5Sclose, _sid),
                             (_lib.H5Tclose, _et), (_lib.H5Tclose, _base),
                             (_lib.H5Fclose, _fid)):
            _close(_obj)
    except Exception as _e:                      # noqa: BLE001
        print(f"warn: could not build tiny.wideenum.h5 via libhdf5 ctypes ({_e}); "
              "keeping any committed copy", file=sys.stderr)

    # tiny.badsparse.h5ad: a HOSTILE AnnData. The CSR `X` group's `shape`
    # attribute lies (claims 100 rows) but `indptr` holds only 2 rows; and the
    # `obs` DataFrame has children of unequal length. vv used to derive row
    # counts / nnz from the untrusted shape (OOB hyperslab read) and to build a
    # table from unequal-length columns (invalid → OOB paging). It must now
    # clamp to the real extents and render a bounded, valid preview — no crash.
    bad_path = HERE / "tiny.badsparse.h5ad"
    if bad_path.exists():
        bad_path.unlink()
    with h5py.File(bad_path, "w") as f:
        f.attrs["encoding-type"] = "anndata"      # marks it as AnnData to vv
        X = f.create_group("X")
        X.attrs["encoding-type"] = "csr_matrix"
        X.attrs["shape"] = np.array([100, 5], dtype=np.int64)   # lies: 2 real rows
        X.create_dataset("indptr",  data=np.array([0, 1, 2], dtype=np.int64))
        X.create_dataset("indices", data=np.array([0, 3], dtype=np.int64))
        X.create_dataset("data",    data=np.array([1.0, 2.0], dtype=np.float64))
        obs = f.create_group("obs")
        obs.attrs["_index"] = "idx"
        obs.create_dataset("idx", data=np.array([0, 1, 2], dtype=np.int64))  # 3
        obs.create_dataset("bad", data=np.array([10, 20], dtype=np.int64))   # 2

    # tiny.nullstr.h5ad: obs/var string columns + the DataFrame _index encoded as
    # `nullable-string-array` groups ({values, mask}) — anndata >= 0.13's on-disk
    # form. Hand-written with h5py so this path is covered regardless of the
    # installed anndata version. `label` has one NA (mask=True) to exercise the
    # mask → null path (which the anndata-generated fixtures don't).
    nsa_path = HERE / "tiny.nullstr.h5ad"
    if nsa_path.exists():
        nsa_path.unlink()
    _sdt = h5py.string_dtype(encoding="utf-8")
    def _nsa(parent, name, values, mask):
        g = parent.create_group(name)
        g.attrs["encoding-type"] = "nullable-string-array"
        g.attrs["encoding-version"] = "0.1.0"
        g.create_dataset("values", data=np.array(values, dtype=object), dtype=_sdt)
        g.create_dataset("mask", data=np.array(mask, dtype=bool))
    with h5py.File(nsa_path, "w") as f:
        f.attrs["encoding-type"] = "anndata"
        f.attrs["encoding-version"] = "0.1.0"
        f.create_dataset("X", data=np.array([[1., 2.], [3., 4.], [5., 6.]],
                                            dtype=np.float32))
        obs = f.create_group("obs")
        obs.attrs["encoding-type"] = "dataframe"
        obs.attrs["encoding-version"] = "0.2.0"
        obs.attrs["_index"] = "_index"
        obs.attrs["column-order"] = np.array(["label"], dtype=object)
        _nsa(obs, "_index", ["r0", "r1", "r2"], [False, False, False])
        _nsa(obs, "label",  ["alpha", "beta", "gamma"], [False, True, False])  # beta→NA
        var = f.create_group("var")
        var.attrs["encoding-type"] = "dataframe"
        var.attrs["encoding-version"] = "0.2.0"
        var.attrs["_index"] = "_index"
        var.attrs["column-order"] = np.array(["gene_name"], dtype=object)
        _nsa(var, "_index", ["g0", "g1"], [False, False])
        _nsa(var, "gene_name", ["GENE0", "GENE1"], [False, False])

try:
    import anndata as ad                                     # type: ignore
    import numpy as np                                       # type: ignore
    from scipy import sparse                                 # type: ignore
except ImportError:
    print("warn: anndata / scipy not found; skipping tiny.h5ad",
          file=sys.stderr)
else:
    h5ad_path = HERE / "tiny.h5ad"
    if h5ad_path.exists():
        h5ad_path.unlink()
    # Tiny dataset: 5 "cells" × 4 "genes" with a CSR-sparse X.
    rng = np.random.default_rng(42)
    rows, cols, vals = [], [], []
    for r in range(5):
        for c in range(4):
            if rng.random() < 0.5:
                rows.append(r); cols.append(c)
                vals.append(float(rng.integers(1, 100)))
    X = sparse.csr_matrix((vals, (rows, cols)), shape=(5, 4),
                            dtype=np.float64)
    obs = ad.AnnData.__module__  # marker for older anndata versions
    import pandas as pd                                       # type: ignore
    obs_df = pd.DataFrame({
        "cluster":  pd.Categorical(["A", "B", "A", "B", "A"]),
        "n_counts": np.array([12, 7, 19, 4, 25], dtype=np.float64),
    }, index=[f"cell{i}" for i in range(5)])
    var_df = pd.DataFrame({
        "gene_name": ["G1", "G2", "G3", "G4"],
        "mt":        [False, False, True, False],
    }, index=[f"gene{i}" for i in range(4)])
    # The three dense-2D shapes are deliberately all present and all distinct,
    # because vv labels their axes differently and used to label them the same:
    #   obsm/X_umap  (5 obs x 2 dims)  -> cells x embedding dimensions
    #   varm/PCs     (4 var x 3 comps) -> GENES x components (rows are var!)
    #   layers/counts(5 obs x 4 var)   -> same shape as X, so genes as columns
    # A width of 2 / 3 / 4 keeps them mutually distinguishable in assertions.
    adata = ad.AnnData(X=X, obs=obs_df, var=var_df,
                          obsm={"X_umap": np.array(
                              [[0.1, 0.2], [-0.3, 0.4],
                               [0.5, -0.1], [-0.2, -0.5],
                               [0.0, 0.0]], dtype=np.float64)},
                          varm={"PCs": np.array(
                              [[1.0, 2.0, 3.0], [4.0, 5.0, 6.0],
                               [7.0, 8.0, 9.0], [10.0, 11.0, 12.0]],
                              dtype=np.float64)},
                          layers={"counts": np.arange(
                              20, dtype=np.float64).reshape(5, 4)})
    adata.write_h5ad(h5ad_path)

    # tiny.legacy_cat.h5ad: the anndata < 0.8 categorical encoding, written by
    # hand because modern anndata cannot emit it. A categorical column is a
    # plain integer code array whose `categories` attribute is an HDF5 object
    # reference to a lookup table parked in a `__categories` group beside the
    # columns — NOT the modern {codes, categories} sub-group. vv read the codes
    # raw, so a real Perturb-seq file showed `gene = 1_157` and `strand = 0`
    # instead of gene names and +/-.
    legacy_path = HERE / "tiny.legacy_cat.h5ad"
    if legacy_path.exists():
        legacy_path.unlink()
    with h5py.File(legacy_path, "w") as h:
        n = 4
        obs = h.create_group("obs")
        obs.attrs["_index"] = "cell_barcode"
        obs.attrs["column-order"] = np.array(["gene", "strand", "n_counts"],
                                             dtype=object)
        obs.attrs["encoding-type"] = "dataframe"
        obs.attrs["encoding-version"] = "0.1.0"
        obs.create_dataset("cell_barcode",
                           data=np.array([f"cell{i}".encode() for i in range(n)]))
        obs.create_dataset("n_counts", data=np.array([10, 20, 30, 40],
                                                     dtype=np.float32))
        cats = obs.create_group("__categories")
        # int16 codes + a 3-entry dictionary; code 2 is used twice and the
        # order is deliberately not sorted, so a wrong decode is visible.
        cats.create_dataset("gene",
                            data=np.array([b"BRCA1", b"TP53", b"EGFR"]))
        g = obs.create_dataset("gene", data=np.array([2, 0, 1, 2],
                                                     dtype=np.int16))
        g.attrs["categories"] = cats["gene"].ref
        # int8 codes, and the classic genomics case: +/- must not read as 0/1.
        cats.create_dataset("strand", data=np.array([b"+", b"-"]))
        st = obs.create_dataset("strand", data=np.array([0, 1, 1, 0],
                                                        dtype=np.int8))
        st.attrs["categories"] = cats["strand"].ref
        var = h.create_group("var")
        var.attrs["_index"] = "gene_id"
        var.attrs["encoding-type"] = "dataframe"
        var.attrs["encoding-version"] = "0.1.0"
        var.create_dataset("gene_id",
                           data=np.array([b"ENSG1", b"ENSG2"]))
        h.create_dataset("X", data=np.arange(n * 2, dtype=np.float32).reshape(n, 2))

    # tiny.dense.h5ad: a *dense* X that is wider than the 200-column dense
    # preview cap (3 cells × 250 genes). scan_anndata emits a Matrix2D tab for
    # a dense X with no column gate, so without the cap this densifies the whole
    # matrix into RAM (the OOM/DoS guarded by read_2d_dataset_table). Here it
    # must preview as 3 × 200 with a "first 200 of 250 cols" footer.
    dense_path = HERE / "tiny.dense.h5ad"
    if dense_path.exists():
        dense_path.unlink()
    n_obs, n_var = 3, 250
    Xd = np.arange(n_obs * n_var, dtype=np.float64).reshape(n_obs, n_var)
    dobs = pd.DataFrame({"cluster": pd.Categorical(["A", "B", "A"])},
                        index=[f"cell{i}" for i in range(n_obs)])
    dvar = pd.DataFrame(index=[f"gene{i}" for i in range(n_var)])
    ad.AnnData(X=Xd, obs=dobs, var=dvar).write_h5ad(dense_path)

    # tiny.bigobs.h5ad: 1500 obs rows — exceeds the 1000-row component preview
    # cap, so the obs / X tabs must render "preview: first 1000 of 1500 rows"
    # instead of reading every row (the fix for stalling on huge AnnData
    # components over a slow mount).
    big_path = HERE / "tiny.bigobs.h5ad"
    if big_path.exists():
        big_path.unlink()
    nbig = 1500
    Xb = sparse.random(nbig, 4, density=0.3, format="csr", dtype=np.float64,
                       random_state=0)
    bobs = pd.DataFrame({
        "grp": pd.Categorical((["A", "B", "C"] * ((nbig // 3) + 1))[:nbig]),
        "val": np.arange(nbig, dtype=np.float64),
    }, index=[f"cell{i}" for i in range(nbig)])
    bvar = pd.DataFrame(index=[f"gene{i}" for i in range(4)])
    ad.AnnData(X=Xb, obs=bobs, var=bvar).write_h5ad(big_path)

    # tiny.uns.h5ad: a populated uns (unstructured) dict — string/int/float
    # scalars, a string array, and a nested dict — to exercise the uns key/value
    # tab. Fixed values so the test can assert them (incl. the dotted key from
    # the nested dict). uns was previously skipped entirely by the reader.
    uns_path = HERE / "tiny.uns.h5ad"
    if uns_path.exists():
        uns_path.unlink()
    aun = ad.AnnData(X=sparse.csr_matrix(np.eye(3, dtype=np.float64)),
                     obs=pd.DataFrame(index=[f"cell{i}" for i in range(3)]),
                     var=pd.DataFrame(index=[f"gene{i}" for i in range(3)]))
    aun.uns["title"]     = "demo dataset"
    aun.uns["n_pcs"]     = 50
    aun.uns["threshold"] = 0.05
    aun.uns["method"]    = "leiden"
    aun.uns["X_colors"]  = np.array(["#FF0000", "#00FF00", "#0000FF"])
    aun.uns["pca"]       = {"variance_ratio": np.array([0.5, 0.3, 0.2])}
    aun.write_h5ad(uns_path)
    # tiny.csc.h5ad: a CSC-sparse X (encoding-type csc_matrix). The preview must
    # densify identically to CSR — rows × columns — by walking each column's
    # indptr range (indices are row indices). Fixed values so the test can
    # assert exact cells; gene/cell identifiers exercise the X-labels path too.
    csc_path = HERE / "tiny.csc.h5ad"
    if csc_path.exists():
        csc_path.unlink()
    dense = np.array([[1., 0., 0., 2.],
                      [0., 3., 0., 0.],
                      [0., 0., 4., 5.]], dtype=np.float64)
    Xc = sparse.csc_matrix(dense)
    cobs = pd.DataFrame({"cluster": pd.Categorical(["A", "B", "A"])},
                        index=[f"cell{i}" for i in range(3)])
    cvar = pd.DataFrame({"gene_name": ["G1", "G2", "G3", "G4"]},
                        index=[f"gene{i}" for i in range(4)])
    ad.AnnData(X=Xc, obs=cobs, var=cvar).write_h5ad(csc_path)

# tiny.malformed.h5ad: a hostile AnnData that previously crashed the reader.
# Derived from the valid tiny.h5ad (so it stays AnnData-detectable and reaches
# the sparse-X path) by poking two attributes that HDF5's H5Aread fills one
# element per dataspace point:
#   - /X "shape": 16 int64 values instead of 2 → smashed the fixed int64[2]
#   - /obs "_index": a 6-element string array → smashed a single-string buffer
# vv must now open it without crashing (it reads only the first two dims and
# treats the array-valued string attr as absent).
try:
    import h5py                                              # type: ignore
    import numpy as np                                       # type: ignore
except ImportError:
    pass
else:
    base = HERE / "tiny.h5ad"
    if base.exists():
        import shutil
        mal = HERE / "tiny.malformed.h5ad"
        if mal.exists():
            mal.unlink()
        shutil.copy(base, mal)
        with h5py.File(mal, "r+") as f:
            if "X" in f:
                if "shape" in f["X"].attrs:
                    del f["X"].attrs["shape"]
                f["X"].attrs.create("shape", np.arange(16, dtype=np.int64))
            if "obs" in f:
                if "_index" in f["obs"].attrs:
                    del f["obs"].attrs["_index"]
                f["obs"].attrs.create(
                    "_index", np.array([b"i0", b"i1", b"i2",
                                        b"i3", b"i4", b"i5"], dtype="S2"))

# ── NumPy .npz fixtures ──────────────────────────────────────────────────────
# tiny.npz: a valid archive with a 1-D, 2-D and 3-D array — exercises the
# NPZ summary tab and the per-array densification paths.
# tiny.malformed.npz: a hostile archive whose single member's .npy header
# declares shape (1000000000,) <f8 (8 GB) but stores only 16 bytes. The reader
# derived element counts / byte offsets straight from the shape, so this drove
# an out-of-bounds read; vv must now reject it at open with a clear error.
try:
    import numpy as np                                       # type: ignore
    import zipfile, struct                                   # type: ignore
except ImportError:
    print("warn: numpy not found; skipping tiny.npz / tiny.malformed.npz",
          file=sys.stderr)
else:
    np.savez(HERE / "tiny.npz",
             vec=np.arange(6, dtype=np.int64),
             mat=np.arange(12, dtype=np.float64).reshape(3, 4),
             cube=np.arange(24, dtype=np.float32).reshape(2, 3, 4))

    # tiny.dtypes.npz: one 1-D array per NumPy dtype vv's slab_to_arrow()
    # can build. arrow_type_for_id() used to map only INT64/DOUBLE/BINARY and
    # fall through to utf8() for the rest, so the declared schema said `string`
    # while the chunk carried the real array. Arrow does not check that on the
    # write path: `--parquet` failed loudly but `--arrow` exited 0 and wrote an
    # IPC file nothing could read back, for 7 of these 9 dtypes.
    np.savez(HERE / "tiny.dtypes.npz",
             i8=np.array([-8, 0, 8], dtype=np.int8),
             i16=np.array([-16, 0, 16], dtype=np.int16),
             i32=np.array([-32, 0, 32], dtype=np.int32),
             i64=np.array([-64, 0, 64], dtype=np.int64),
             u8=np.array([0, 8, 255], dtype=np.uint8),
             u16=np.array([0, 16, 65535], dtype=np.uint16),
             u32=np.array([0, 32, 4294967295], dtype=np.uint32),
             u64=np.array([0, 64, 18446744073709551615], dtype=np.uint64),
             f32=np.array([-1.5, 0.0, 1.5], dtype=np.float32),
             f64=np.array([-2.5, 0.0, 2.5], dtype=np.float64),
             b=np.array([True, False, True], dtype=bool))

    # tiny.zerorow.npz: a Fortran-ordered 2-D array with zero rows. The
    # per-column gather sized its scratch buffer to rows * item_size, so at
    # zero rows the buffer was empty and its data() null — and the F-order
    # branch then called memcpy(nullptr, ..., 0), which is undefined even at
    # zero length. Found by the .npy fuzzer; this fixture pins it so the
    # ASan/UBSan CI job catches a regression without relying on the fuzzer
    # reaching the same input again.
    np.savez(HERE / "tiny.zerorow.npz",
             empty_f=np.asfortranarray(np.zeros((0, 3), dtype=np.float64)),
             empty_c=np.zeros((0, 3), dtype=np.float64))

    # tiny.dupmember.npz: a HOSTILE archive with two members both named
    # "arr.npy" — shapes (3,) then (5,). numpy's np.load keeps the last (its
    # central-directory dict overwrites earlier keys); vv used to keep the
    # first, so it displayed a different array than numpy and left the second
    # as an unreachable, shadowed tab. vv now keeps the last, matching numpy,
    # and warns. numpy's savez refuses duplicate keys, so write the ZIP by hand.
    import io, warnings                                       # type: ignore
    def _npy_bytes(a):
        b = io.BytesIO(); np.save(b, a); return b.getvalue()
    with zipfile.ZipFile(HERE / "tiny.dupmember.npz", "w") as _z, \
         warnings.catch_warnings():
        warnings.simplefilter("ignore")                      # dup-name UserWarning
        _z.writestr("arr.npy",  _npy_bytes(np.array([1, 2, 3], dtype=np.int64)))
        _z.writestr("keep.npy", _npy_bytes(np.array([7, 7], dtype=np.int64)))
        _z.writestr("arr.npy",  _npy_bytes(np.arange(100, 105, dtype=np.int64)))

    # tiny.shapeovf.npz: a shape that declares an EMPTY array (trailing 0) while
    # its leading dimensions multiply past int64. One zero dim makes the total
    # element count zero, so the size check passes for free — but the 3-D+
    # reader collapses the trailing dims into a stride and overflowed doing it.
    # Found by the .npy fuzzer; reachable just by opening the file. numpy won't
    # write such a header, so build it by hand.
    _ovf_hdr = ("{'descr': '|b1', 'fortran_order': False, "
                "'shape': (1, 1, 392361265078550784, 29, 0), }")
    _ovf_hdr += " " * ((64 - (10 + len(_ovf_hdr) + 1)) % 64) + "\n"
    _ovf_npy = (b"\x93NUMPY\x01\x00" + struct.pack("<H", len(_ovf_hdr))
                + _ovf_hdr.encode())
    _ovf = HERE / "tiny.shapeovf.npz"
    if _ovf.exists():
        _ovf.unlink()
    with zipfile.ZipFile(_ovf, "w", zipfile.ZIP_STORED) as z:
        z.writestr("a.npy", _ovf_npy)

    # tiny.wide.npz: a 2-D array wider than the NPZ column cap
    # (kNpzMaxCols = 4096). vv builds one Arrow column per declared column,
    # so a genuinely-wide (or hostile) array would otherwise allocate
    # unboundedly; vv must render only the first 4096 columns and flag the
    # truncation in the footer.
    np.savez(HERE / "tiny.wide.npz",
             wide=np.arange(2 * 5000, dtype=np.int32).reshape(2, 5000))

    def _make_npy(descr, shape, data):
        tup = "(" + ", ".join(str(s) for s in shape) + \
              ("," if len(shape) == 1 else "") + ")"
        hdr = "{'descr': '%s', 'fortran_order': False, 'shape': %s, }" % (descr, tup)
        pad = (64 - (10 + len(hdr) + 1) % 64) % 64
        hdr = hdr + " " * pad + "\n"
        return (b"\x93NUMPY\x01\x00" + struct.pack("<H", len(hdr))
                + hdr.encode() + data)

    mal_npz = HERE / "tiny.malformed.npz"
    if mal_npz.exists():
        mal_npz.unlink()
    with zipfile.ZipFile(mal_npz, "w", zipfile.ZIP_STORED) as z:
        # header claims 1e9 float64 elements; body is only 16 bytes
        z.writestr("a.npy", _make_npy("<f8", (1000000000,), b"\x00" * 16))

    # tiny.badsize.npz: a VALID array, but the central-directory
    # `uncompressed_size` field — which is attacker-controllable and was passed
    # straight to vector::reserve() — is patched to a bogus ~2 GiB. The body is
    # still read from the real stored stream, so vv must read the array
    # correctly without trusting the size hint for allocation.
    bad_size_npz = HERE / "tiny.badsize.npz"
    if bad_size_npz.exists():
        bad_size_npz.unlink()
    np.savez(bad_size_npz, arr=np.arange(5, dtype=np.int64))
    raw = bytearray(bad_size_npz.read_bytes())
    cd = raw.find(b"\x50\x4b\x01\x02")          # central-directory file header
    if cd >= 0:                                  # uncompressed_size: +24 past sig
        struct.pack_into("<I", raw, cd + 24, 0x7FFFFFFF)
        bad_size_npz.write_bytes(raw)

# ── Apache ORC (columnar; pyarrow.orc is built into pyarrow when Arrow was
# compiled with -DARROW_ORC=ON, which is the wheel default since pyarrow
# 12.0). Soft-skip otherwise.
try:
    import pyarrow.orc as paorc                            # type: ignore
except ImportError:
    print("warn: pyarrow.orc not available; skipping tiny.orc", file=sys.stderr)
else:
    orc_path = HERE / "tiny.orc"
    if orc_path.exists():
        orc_path.unlink()
    orc_tbl = pa.table({
        "chrom": pa.array(["chr1", "chr1", "chr2"], type=pa.string()),
        "start": pa.array([100, 500, 1000], type=pa.int64()),
        "end":   pa.array([200, 800, 1300], type=pa.int64()),
        "score": pa.array([5.2, 8.1, 3.4],  type=pa.float64()),
        "name":  pa.array(["p1", "p2", "p3"], type=pa.string()),
    })
    # 2 rows per stripe → 2 stripes so num_chunks > 1 in vv's output.
    paorc.write_table(orc_tbl, orc_path, compression="zstd", stripe_size=2)

# ── Excel (.xlsx, two sheets, requires openpyxl) ─────────────────────────────
# Soft-skip if openpyxl isn't installed (CI installs it in the venv alongside
# pyarrow; local dev can `pip install openpyxl` to regenerate).
try:
    import openpyxl                                        # type: ignore
except ImportError:
    print("warn: openpyxl not found; skipping tiny.xlsx", file=sys.stderr)
else:
    xlsx_path = HERE / "tiny.xlsx"
    if xlsx_path.exists():
        xlsx_path.unlink()
    wb = openpyxl.Workbook()
    # First sheet — replace the default name to assert sheet-naming works.
    s1 = wb.active
    assert s1 is not None
    s1.title = "peaks"
    s1.append(["chrom", "start", "end", "score", "name"])
    s1.append(["chr1", 100, 200, 5.2, "p1"])
    s1.append(["chr1", 500, 800, 8.1, "p2"])
    s1.append(["chr2", 1000, 1300, 3.4, "p3"])
    # Second sheet — exercises sibling-tab expansion in main().
    s2 = wb.create_sheet("samples")
    s2.append(["sample_id", "name", "depth"])
    s2.append([1, "sampleA", 12.5])
    s2.append([2, "sampleB", 8.7])
    s2.append([3, "sampleC", 15.1])
    wb.save(xlsx_path)

    # Ragged workbook: a row wider than the header (.xlsx). See the matching
    # tiny.ragged.ods note — the wider row used to make Arrow reject the sheet.
    rg_xlsx = HERE / "tiny.ragged.xlsx"
    if rg_xlsx.exists():
        rg_xlsx.unlink()
    rwb = openpyxl.Workbook()
    rs = rwb.active
    assert rs is not None
    rs.title = "ragged"
    rs.append(["chrom", "start", "end"])
    rs.append(["chr1", 100, 200])
    rs.append(["chr2", 300, 400, "EXTRA"])   # wider than the header
    rs.append(["chr3", 500])                  # shorter than the header
    rwb.save(rg_xlsx)

# ── BCF (binary VCF, requires bcftools) ──────────────────────────────────────
if have("bcftools"):
    # Make a self-contained VCF (with explicit ##contig lines) so bcftools is
    # happy; tiny.vcf above lacks those.
    bcf_input = HERE / "_bcf_input.vcf"
    bcf_input.write_text(
        "##fileformat=VCFv4.2\n"
        "##contig=<ID=chr1>\n"
        "##contig=<ID=chr2>\n"
        '##INFO=<ID=AF,Number=A,Type=Float,Description="Allele frequency">\n'
        "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\n"
        "chr1\t100\trs1\tA\tG\t30\tPASS\tAF=0.5\n"
        "chr1\t500\t.\tC\tT\t40\tPASS\tAF=0.1\n"
        "chr1\t1500\t.\tG\tA\t50\tPASS\tAF=0.3\n"
        "chr2\t200\t.\tT\tC\t35\tPASS\tAF=0.2\n"
    )
    subprocess.run(["bcftools", "view", "-O", "b",
                    str(bcf_input), "-o", str(HERE / "tiny.bcf")], check=True)
    subprocess.run(["bcftools", "index", "-f",
                    str(HERE / "tiny.bcf")], check=True)
    bcf_input.unlink()

    # tiny.samples.bcf: a BCF WITH genotype samples, so the reader emits a
    # FORMAT_SAMPLES column. Exercises that the FORMAT spec (GT:AD:DP) is kept,
    # not dropped in favour of the sample columns alone.
    smp_input = HERE / "_bcf_samples.vcf"
    smp_input.write_text(
        "##fileformat=VCFv4.2\n"
        "##contig=<ID=chr1>\n"
        '##INFO=<ID=AF,Number=A,Type=Float,Description="Allele frequency">\n'
        '##FORMAT=<ID=GT,Number=1,Type=String,Description="Genotype">\n'
        '##FORMAT=<ID=AD,Number=R,Type=Integer,Description="Allelic depths">\n'
        '##FORMAT=<ID=DP,Number=1,Type=Integer,Description="Read depth">\n'
        "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tS1\tS2\n"
        "chr1\t100\trs1\tA\tG\t30\tPASS\tAF=0.5\tGT:AD:DP\t0/1:5,6:11\t1/1:0,9:9\n"
        "chr1\t500\t.\tC\tT\t40\tPASS\tAF=0.1\tGT:AD:DP\t0/0:10,0:10\t0/1:4,4:8\n"
    )
    subprocess.run(["bcftools", "view", "-O", "b",
                    str(smp_input), "-o", str(HERE / "tiny.samples.bcf")],
                   check=True)
    subprocess.run(["bcftools", "index", "-f",
                    str(HERE / "tiny.samples.bcf")], check=True)
    smp_input.unlink()
else:
    print("warn: bcftools not found; skipping tiny.bcf", file=sys.stderr)

# ── bigBed / bigWig ──────────────────────────────────────────────────────────
# Need UCSC's bedToBigBed / bedGraphToBigWig. They're in the
# `ucsc-bedtobigbed` and `ucsc-bedgraphtobigwig` Bioconda packages, on
# Arch under `kentutils`, in /opt/ucsc-kent-genome-tools on this dev box.
def find_kent_tool(name):
    if shutil.which(name): return name
    cand = "/opt/ucsc-kent-genome-tools/" + name
    if (HERE.parent.parent.parent / cand[1:]).exists() or shutil.which(cand): return cand
    import os
    if os.path.exists(cand): return cand
    return None

bb_tool = find_kent_tool("bedToBigBed")
bw_tool = find_kent_tool("bedGraphToBigWig")
if bb_tool and bw_tool:
    sizes = HERE / "_sizes.txt"
    sizes.write_text("chr1\t50000\nchr2\t50000\n")
    # bigBed with BED6+3 autoSql so the autoSql parser is actually exercised.
    bb_bed = HERE / "_peaks.bed"
    bb_bed.write_text(
        "chr1\t100\t200\tpeak_0\t500\t+\t12.5\t30.2\t25.1\n"
        "chr1\t500\t800\tpeak_1\t800\t-\t8.0\t15.0\t12.0\n"
        "chr1\t1000\t1200\tpeak_2\t300\t+\t5.5\t8.1\t7.0\n"
        "chr2\t200\t400\tpeak_3\t950\t.\t20.1\t40.5\t35.2\n"
        "chr2\t1500\t1800\tpeak_4\t150\t+\t2.0\t3.0\t2.5\n"
    )
    bb_as = HERE / "_peaks.as"
    bb_as.write_text(
        'table bigBed6Plus3\n'
        '"BED6 + 3 narrow-peak columns"\n'
        '(\n'
        '    string chrom;        "Reference sequence"\n'
        '    uint   chromStart;   "Start"\n'
        '    uint   chromEnd;     "End"\n'
        '    string name;         "Item name"\n'
        '    uint   score;        "0..1000"\n'
        '    char[1] strand;      "+/-/."\n'
        '    float  signalValue;  "Enrichment"\n'
        '    float  pValue;       "-log10 p"\n'
        '    float  qValue;       "-log10 q"\n'
        ')\n')
    subprocess.run([bb_tool, "-type=bed6+3", "-as=" + str(bb_as),
                    str(bb_bed), str(sizes), str(HERE / "tiny.bb")], check=True)
    # bigWig from a bedGraph.
    bg = HERE / "_signal.bedGraph"
    bg.write_text(
        "chr1\t100\t200\t0.5\n"
        "chr1\t500\t800\t0.7\n"
        "chr1\t1000\t1200\t0.2\n"
        "chr2\t200\t400\t0.9\n"
        "chr2\t1500\t1800\t0.1\n"
    )
    subprocess.run([bw_tool, str(bg), str(sizes), str(HERE / "tiny.bw")], check=True)

    # tiny.badnitems.bw: HOSTILE bigWig. An uncompressed data block's header
    # declares nItems=65535 while the block holds only four intervals. libBigWig
    # trusted nItems and read 24 + 65535*12 bytes from a 72-byte block — an
    # out-of-bounds heap read (glibc/ASan abort). vv now bounds each record to
    # the block, reading the four real intervals and stopping. Built with -unc so
    # the nItems field can be patched in place without recompression.
    nbg = HERE / "_nitems.bedGraph"
    nbg.write_text("chr1\t0\t10\t1.0\nchr1\t10\t20\t2.0\n"
                   "chr1\t20\t30\t3.0\nchr1\t30\t40\t4.0\n")
    subprocess.run([bw_tool, "-unc", str(nbg), str(sizes),
                    str(HERE / "tiny.badnitems.bw")], check=True)
    _nb = bytearray((HERE / "tiny.badnitems.bw").read_bytes())
    _ndo = struct.unpack_from("<Q", _nb, 0x10)[0]        # dataOffset
    struct.pack_into("<H", _nb, _ndo + 8 + 22, 65535)    # inflate the block nItems
    (HERE / "tiny.badnitems.bw").write_bytes(bytes(_nb))
    nbg.unlink()

    # tiny.unterminated.bb: HOSTILE bigBed. The final entry's name string has its
    # NUL terminator overwritten, so it runs to the end of the uncompressed data
    # block. libBigWig's strlen then read past the buffer (out-of-bounds heap
    # read, ASan abort). vv now bounds the string to the block and stops at the
    # unterminated entry. Built with -unc so the edit needs no recompression.
    ubed = HERE / "_unterm.bed"
    ubed.write_text("chr1\t100\t200\tpeakAAAA\t500\t+\t12.5\t30.2\t25.1\n"
                    "chr1\t500\t800\tpeakBBBB\t800\t-\t8.0\t15.0\t12.0\n"
                    "chr1\t1000\t1200\tpeakCCCC\t300\t+\t5.5\t8.1\t7.0\n")
    subprocess.run([bb_tool, "-unc", "-type=bed6+3", "-as=" + str(bb_as),
                    str(ubed), str(sizes), str(HERE / "tiny.unterminated.bb")],
                   check=True)
    _ub = bytearray((HERE / "tiny.unterminated.bb").read_bytes())
    _uio = struct.unpack_from("<Q", _ub, 0x18)[0]        # indexOffset = data end
    if _ub[_uio - 1] == 0:                               # final block byte is a NUL
        _ub[_uio - 1] = ord("X")
        (HERE / "tiny.unterminated.bb").write_bytes(bytes(_ub))
    else:
        print("warn: tiny.unterminated.bb final byte not NUL; skipping edit",
              file=sys.stderr)
    ubed.unlink()

    sizes.unlink(); bb_bed.unlink(); bb_as.unlink(); bg.unlink()
else:
    print("warn: bedToBigBed / bedGraphToBigWig not found; "
          "skipping tiny.bb and tiny.bw", file=sys.stderr)

# tiny.oobchrom.bw: a HOSTILE bigWig. The chromosome B-tree leaf names a
# chromId (64) outside the itemCount the same file declares (2). libBigWig used
# that value directly as the subscript for writes into cl->len and cl->chrom,
# both calloc'd to itemCount entries, so opening the file wrote four
# file-chosen bytes and a pointer past the end of two heap allocations —
# glibc aborted on the corrupted heap; ASan reports heap-buffer-overflow.
# vv must now reject the file instead.
#
# Derived by patching the committed tiny.bw rather than built from scratch, so
# it is reproducible without UCSC's kent tools (absent on macOS CI and most
# dev boxes, which is why tiny.bw itself is gated above).
bw_src = HERE / "tiny.bw"
if bw_src.exists():
    _d = bytearray(bw_src.read_bytes())
    _ct = struct.unpack_from("<Q", _d, 8)[0]            # chromosomeTreeOffset
    _keySize = struct.unpack_from("<I", _d, _ct + 8)[0]
    _node = _ct + 32                                     # past the cirTree header
    _isLeaf, _pad, _nVals = struct.unpack_from("<BBH", _d, _node)
    if _isLeaf and _nVals:
        # item layout: key[keySize], chromId(u32), chromSize(u32)
        struct.pack_into("<I", _d, _node + 4 + _keySize, 64)
        (HERE / "tiny.oobchrom.bw").write_bytes(bytes(_d))
    else:
        print("warn: tiny.bw chrom tree is not a populated leaf; "
              "skipping tiny.oobchrom.bw", file=sys.stderr)
else:
    print("warn: tiny.bw absent; skipping tiny.oobchrom.bw", file=sys.stderr)

# ── TSV / CSV ────────────────────────────────────────────────────────────────
tsv = "name\tcount\tratio\nfoo\t100\t0.5\nbar\t250\t0.75\nbaz\t9999\t0.1\n"
(HERE / "tiny.tsv").write_text(tsv)
csv = tsv.replace("\t", ",")
(HERE / "tiny.csv").write_text(csv)

print("done; files in", HERE)
